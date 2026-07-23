"""utils/output_xlsx.py — xlsx 清单输出

从 output_format.py 按介质拆分而来（xlsx 部分）。公开函数
export_to_excel_with_format / export_to_one_excel_with_format / export_to_excel_twoheader。
既有代码经 output_format.py 聚合入口导入即可，无需改动。
"""
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


def _char_width(ch):
    """估算单个字符在 Excel 中的列宽。CJK 字符约占 2 单位，其余占 1 单位。"""
    code = ord(ch)
    if (0x4E00 <= code <= 0x9FFF or 0x3400 <= code <= 0x4DBF
            or 0xF900 <= code <= 0xFAFF or 0x2E80 <= code <= 0x2FDF
            or 0x3000 <= code <= 0x303F or 0xFF01 <= code <= 0xFF60
            or 0xFE30 <= code <= 0xFE4F or 0x2000 <= code <= 0x206F):
        return 2.0
    return 1.0


def _col_widths(df):
    """计算 DataFrame 各列自适应宽度（基于前 100 行采样）。"""
    sample = df.head(100)
    widths = []
    for c in df.columns:
        header_w = sum(_char_width(ch) for ch in str(c))
        max_w = header_w
        for val in sample[c]:
            if pd.isna(val):
                continue
            w = sum(_char_width(ch) for ch in str(val))
            if w > max_w:
                max_w = w
        widths.append(min(max_w + 2, 60))
    return widths


def _sanitize_sheet_name(name: str) -> str:
    """清理 Excel sheet name 中的全角冒号（Excel 不支持 `：`）。"""
    return name.replace("：", "-")


# ── export_to_excel_with_format (xlsxwriter) ──────────────────────────


def export_to_excel_with_format(df, output_path, sheet_name, title_name, add_title=True):
    """将 DataFrame 输出为格式化的 Excel 清单（xlsxwriter）。"""
    sheet_name = _sanitize_sheet_name(sheet_name)
    num_cols = df.shape[1]
    num_rows = df.shape[0]
    header_row = 1 if add_title else 0
    data_start_row = header_row + 1
    col_widths_list = _col_widths(df)
    text_cols = {i for i, c in enumerate(df.columns) if df[c].dtype == object}

    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        workbook = writer.book
        worksheet = writer.sheets[sheet_name] if sheet_name in writer.sheets \
            else workbook.add_worksheet(sheet_name)

        fmt_header = workbook.add_format({
            'bold': True, 'align': 'center', 'valign': 'vcenter',
            'border': 1, 'bg_color': '#D3D3D3', 'font_size': 10,
        })
        fmt_title = workbook.add_format({
            'bold': True, 'align': 'center', 'valign': 'vcenter',
            'font_size': 14,
        })
        fmt_data = workbook.add_format({
            'border': 1, 'valign': 'vcenter', 'align': 'left',
            'font_size': 10,
        })
        fmt_text = workbook.add_format({
            'border': 1, 'valign': 'vcenter', 'align': 'left',
            'font_size': 10, 'num_format': '@',
        })

        if add_title:
            worksheet.merge_range(0, 0, 0, num_cols - 1, title_name, fmt_title)

        # 表头
        for c, col_name in enumerate(df.columns):
            worksheet.write(header_row, c, col_name, fmt_header)

        # 数据
        for r in range(num_rows):
            for c in range(num_cols):
                val = df.iloc[r, c]
                if pd.isna(val):
                    val = ''
                fmt = fmt_text if c in text_cols else fmt_data
                worksheet.write(data_start_row + r, c, val, fmt)

        # 列宽
        for c, width in enumerate(col_widths_list):
            worksheet.set_column(c, c, width)

        # 筛选
        if num_rows > 0:
            worksheet.autofilter(header_row, 0, data_start_row + num_rows - 1, num_cols - 1)

    print(f"Sheet '{sheet_name}' 已成功导出至: {output_path}")


# ── export_to_one_excel_with_format (openpyxl) ────────────────────────


def export_to_one_excel_with_format(df, output_path, sheet_name, title_name=None, add_title=True):
    """向指定文件写入一个 sheet（已存在则覆盖，否则新建），使用 openpyxl。"""
    sheet_name = _sanitize_sheet_name(sheet_name)
    if os.path.exists(output_path):
        workbook = load_workbook(output_path)
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        worksheet = workbook.create_sheet(sheet_name)
    else:
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name

    num_cols = df.shape[1]
    rows, cols = df.shape
    col_widths_list = _col_widths(df)
    text_cols = {i for i, c in enumerate(df.columns) if df[c].dtype == object}

    # 样式
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    title_font = Font(bold=True, size=14)
    title_alignment = Alignment(horizontal="center", vertical="center")
    data_font = Font(size=10)
    data_alignment = Alignment(vertical="center")
    data_border = Border(
        left=Side(border_style="thin"), right=Side(border_style="thin"),
        top=Side(border_style="thin"), bottom=Side(border_style="thin"),
    )

    # 大标题
    header_row = 2 if add_title else 1
    if add_title and title_name:
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        worksheet["A1"].value = title_name
        worksheet["A1"].font = title_font
        worksheet["A1"].alignment = title_alignment

    # 表头
    for c, col_name in enumerate(df.columns):
        cell = worksheet.cell(row=header_row, column=c + 1)
        cell.value = col_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 数据
    data_start = header_row + 1
    for r in range(rows):
        for c in range(cols):
            val = df.iloc[r, c]
            if pd.isna(val):
                val = ""
            cell = worksheet.cell(row=data_start + r, column=c + 1)
            cell.value = val
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = data_border
            if c in text_cols:
                cell.number_format = '@'

    # 列宽
    for c, width in enumerate(col_widths_list):
        worksheet.column_dimensions[get_column_letter(c + 1)].width = width

    # 筛选
    if rows >= 1 and cols >= 1:
        last_col = get_column_letter(cols)
        worksheet.auto_filter.ref = f"A{header_row}:{last_col}{data_start + rows - 1}"

    workbook.save(output_path)
    print(f"Sheet '{sheet_name}' 已成功导出至: {output_path}")


# ── export_to_excel_twoheader (xlsxwriter) ────────────────────────────


def export_to_excel_twoheader(df, output_path, sheet_name, title,
                              fixed_cols, header_groups,
                              trailing_cols=None, col_widths=None,
                              subject_col=None, count_suffix=None):
    """导出带两层合并表头的 Excel 清单。

    适用于"用药后异常有临床意义"系列清单：固定列 + 分组表头（首次用药前/后）+ 尾列。

    Parameters
    ----------
    df : DataFrame
        输出数据（不含 DataFrame 级表头，列序需与 schema 匹配）。
    output_path : str
        输出文件路径。
    sheet_name : str
        工作表名称。
    title : str
        标题前缀。默认按计数自动追加中文量词（DMR 约定："(N例次M例)" 或 "(N条)"）；
        其他场景/语言用 ``count_suffix`` 覆盖或置空。
    fixed_cols : list[str]
        跨两行合并的固定列名。
    header_groups : list[dict]
        分组表头定义，每项 ``{'label': str, 'children': list[str]}``。
    trailing_cols : list[str], optional
        尾列名（跨两行合并），默认无。
    col_widths : list[tuple], optional
        列宽定义，每项 ``(start_col, end_col, width)``。
    subject_col : str, optional
        用于计算唯一例数的列名，标题中显示 "(N例次M例)"。
    count_suffix : str, optional
        自定义标题计数后缀（覆盖默认中文量词）；传 "" 则标题不加任何后缀。
    """
    sheet_name = _sanitize_sheet_name(sheet_name)
    total_cols = len(fixed_cols) \
               + sum(len(g['children']) for g in header_groups) \
               + len(trailing_cols or [])
    assert total_cols == df.shape[1], \
        f'列数不匹配: schema={total_cols}, df={df.shape[1]}'

    n_rows = len(df)
    n_subj = df[subject_col].nunique() if subject_col else None
    text_cols = {i for i, c in enumerate(df.columns) if df[c].dtype == object}

    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name=sheet_name, startrow=3,
                    index=False, header=False)
        wb = writer.book
        ws = writer.sheets[sheet_name]

        hdr_fmt = wb.add_format({
            'bold': True, 'align': 'center', 'valign': 'vcenter',
            'border': 1, 'bg_color': '#D3D3D3', 'font_size': 10,
        })
        title_fmt = wb.add_format({
            'bold': True, 'align': 'center', 'valign': 'vcenter',
            'font_size': 14,
        })
        data_fmt = wb.add_format({
            'border': 1, 'valign': 'vcenter', 'font_size': 10,
        })
        text_fmt = wb.add_format({
            'border': 1, 'valign': 'vcenter', 'font_size': 10,
            'num_format': '@',
        })

        # 标题
        if count_suffix is not None:
            title_text = f'{title} {count_suffix}'.rstrip()
        elif n_subj is not None:
            title_text = f'{title} ({n_rows}例次{n_subj}例)'
        else:
            title_text = f'{title} ({n_rows}条)'
        ws.merge_range(0, 0, 0, total_cols - 1, title_text, title_fmt)

        # 固定列（跨两行）
        for i, name in enumerate(fixed_cols):
            ws.merge_range(1, i, 2, i, name, hdr_fmt)

        # 分组表头（parent 合并 row 1，children 写 row 2）
        col = len(fixed_cols)
        for group in header_groups:
            n_ch = len(group['children'])
            ws.merge_range(1, col, 1, col + n_ch - 1,
                           group['label'], hdr_fmt)
            for j, child in enumerate(group['children']):
                ws.write(2, col + j, child, hdr_fmt)
            col += n_ch

        # 尾列（跨两行）
        for name in (trailing_cols or []):
            ws.merge_range(1, col, 2, col, name, hdr_fmt)
            col += 1

        # 列宽
        if col_widths:
            for start, end, width in col_widths:
                ws.set_column(start, end, width)
        else:
            col_widths_list = _col_widths(df)
            for c, width in enumerate(col_widths_list):
                ws.set_column(c, c, width)

        # 数据区
        for r in range(n_rows):
            for c in range(total_cols):
                val = df.iloc[r, c]
                if pd.isna(val):
                    val = ''
                fmt = text_fmt if c in text_cols else data_fmt
                ws.write(r + 3, c, val, fmt)

    print(f'{title}已保存为 \'{output_path}\'')
